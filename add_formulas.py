# -*- coding: utf-8 -*-
from openpyxl import load_workbook

file_path = "YS費用明細表_改進版.xlsx"
wb = load_workbook(file_path)

# Sheet 3 is Taiwan payment sheet
ws = wb[wb.sheetnames[3]]
print("Processing Taiwan payment sheet...")

# First, ensure N1 has the exchange rate
if ws['N1'].value is None:
    ws['N1'].value = 32
    print("Set N1 to 32 (exchange rate)")
else:
    print(f"N1 already has value: {ws['N1'].value}")

# Add formulas to first section (rows 6-9) in column F
print("\nAdding formulas to column F (first section):")
for row in range(6, 10):
    cell_f = ws[f'F{row}']
    cell_e = ws[f'E{row}']
    current_val = cell_f.value
    if isinstance(current_val, (int, float)):
        formula = f"=E{row}/$N$1"
        cell_f.value = formula
        print(f"  F{row}: Changed from {current_val} to formula {formula}")
    else:
        print(f"  F{row}: Skipped (not a numeric value: {current_val})")

# Check and add formulas to second section (rows 12+) in column G
print("\nChecking second section (column G):")
for row in range(12, 20):
    cell_g = ws[f'G{row}']
    cell_f = ws[f'F{row}']
    if cell_g.value is not None:
        print(f"  G{row} current value: {cell_g.value}")
        # Check if it's a formula or value
        if isinstance(cell_g.value, str) and cell_g.value.startswith('='):
            print(f"    -> Already has formula")
        elif isinstance(cell_g.value, (int, float)):
            formula = f"=F{row}/$N$1"
            cell_g.value = formula
            print(f"    -> Changed to formula {formula}")

# Save the file
wb.save(file_path)
print("\nFile saved successfully!")

# Verify changes
print("\nVerification:")
print(f"N1: {ws['N1'].value}")
print(f"F6 formula: {ws['F6'].value}")
print(f"F7 formula: {ws['F7'].value}")
print(f"G12 formula: {ws['G12'].value}")

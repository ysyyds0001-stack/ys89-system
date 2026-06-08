# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from shutil import copy2
import os

file_path = "YS費用明細表_改進版.xlsx"

# Create backup first
backup_path = "YS費用明細表_改進版_backup.xlsx"
if os.path.exists(file_path):
    try:
        copy2(file_path, backup_path)
        print(f"Backup created: {backup_path}")
    except Exception as e:
        print(f"Backup failed: {e}")

try:
    # Load file
    wb = load_workbook(file_path, data_only=False)
    print(f"File loaded with {len(wb.sheetnames)} sheets")

    # Check Taiwan payment sheet
    ws = wb[wb.sheetnames[3]]

    # Verify N1
    if ws['N1'].value != 32:
        ws['N1'].value = 32
        print("Updated N1 to 32")
    else:
        print(f"N1 already has value: {ws['N1'].value}")

    # Save file
    wb.save(file_path)
    print("File repaired and saved successfully!")
    print("\nChanges verified:")
    print(f"  - N1 (exchange rate): {ws['N1'].value}")
    print(f"  - F6 formula: {ws['F6'].value}")
    print(f"  - G12 formula: {ws['G12'].value}")

except Exception as e:
    print(f"Error during repair: {e}")
    import traceback
    traceback.print_exc()

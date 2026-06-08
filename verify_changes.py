# -*- coding: utf-8 -*-
from openpyxl import load_workbook

file_path = "YS費用明細表_改進版.xlsx"
wb = load_workbook(file_path)

# Check Taiwan payment sheet (sheet 3)
ws = wb[wb.sheetnames[3]]
print("=== Taiwan Payment Sheet Verification ===")
print(f"N1 (Exchange rate): {ws['N1'].value}")
print()

print("First Section (rows 6-9):")
for row in range(6, 10):
    print(f"  Row {row}:")
    print(f"    E{row} (TWD): {ws[f'E{row}'].value}")
    print(f"    F{row} (USDT formula): {ws[f'F{row}'].value}")

print("\nSecond Section (rows 12-13):")
for row in range(12, 14):
    print(f"  Row {row}:")
    print(f"    D{row} (TWD): {ws[f'D{row}'].value}")
    print(f"    E{row} (Formula check): {ws[f'E{row}'].value}")
    print(f"    F{row} (value): {ws[f'F{row}'].value}")
    print(f"    G{row} (USDT formula): {ws[f'G{row}'].value}")

print("\n=== Summary ===")
print("Taiwan Payment Sheet (1/3 improvements):")
print("  - N1 contains exchange rate: YES" if ws['N1'].value == 32 else "  - N1 exchange rate: NO")
print("  - Formulas in column F: YES" if "=" in str(ws['F6'].value) else "  - Formulas in column F: PARTIAL")
print("  - Formulas in column G: YES" if "=" in str(ws['G12'].value) else "  - Formulas in column G: PARTIAL")
